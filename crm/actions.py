import openpyxl
from io import BytesIO
from django.http import HttpResponse
from django.core.files.storage import default_storage
from .models import Client, ClientProfil, Telephone, AbonnementType, Version, Abonnement, Renouvellement, BoostService, ClientService


def export_to_excel(modeladmin, request, queryset):
    # Create a workbook and add sheets
    workbook = openpyxl.Workbook()
    
    # Create the Client worksheet
    ws_client = workbook.active
    ws_client.title = "Client"
    ws_client.append([
        'Client/Prospect','Nom', 'Prenom', 'Commercial Name', 'Specialties', 'Wilaya', 'Commune',
        'GPS Insertion', 'GPS Insertion Comment', 'Photos Integration', 'Photos Integration Comment',
        'Services Filling', 'Services Filling Comment', 'Client Remark', 'Agenda', 'Agenda Comment',
        'Contract Password', 'Contract Password Comment', 'Premium Search Logo', 'Premium Search Logo Comment',
        'Ad Logo Order', 'Welcome Mail', 'Welcome Mail Comment', 'Activation Call', 'Activation Call Comment',
        'Telephones'
    ])
    
    for client in Client.objects.all():
        try:
            profil = ClientProfil.objects.get(client=client)
        except ClientProfil.DoesNotExist:
            profil = None
        
        telephones = client.telephones.values_list('telephone_number', flat=True)
        phone_numbers = ', '.join(telephones)  # Join phone numbers into a single string

        ws_client.append([
            client.client_or_prospect, client.nom, client.prenom, client.nom_commercial, client.specialites, 
            client.wilaya, client.commune,
            profil.insertion_gps if profil else '', profil.comment_insertion_gps if profil else '', 
            profil.integration_photos if profil else '', profil.comment_integration_photos if profil else '', 
            profil.remplissage_services_cabinet if profil else '', profil.comment_remplissage_services_cabinet if profil else '',
            profil.remarque_client if profil else '', profil.agenda if profil else '', profil.comment_agenda if profil else '',
            profil.mot_de_passe_contrat if profil else '', profil.comment_mot_de_passe_contrat if profil else '', 
            profil.logo_recherche_premium if profil else '', profil.comment_logo_recherche_premium if profil else '', 
            profil.ordre_logo_pub_recherche if profil else '', profil.mail_bienvenue if profil else '', 
            profil.comment_mail_bienvenue if profil else '', profil.appel_activation if profil else '', 
            profil.comment_appel_activation if profil else '', phone_numbers
        ])
    
    # Create the AbonnementType worksheet
    ws_abonnement_type = workbook.create_sheet(title="AbonnementType")
    ws_abonnement_type.append([
        'Abonnement Type', 'Price', 'Version'
    ])
    
    for abonnement_type in AbonnementType.objects.all():
        versions = abonnement_type.versions.values_list('version', flat=True)
        for version in versions:
            ws_abonnement_type.append([
                abonnement_type.abonnement, abonnement_type.prix, version
            ])
    
    # Create the Abonnement worksheet
    ws_abonnement = workbook.create_sheet(title="Abonnement")
    ws_abonnement.append([
        'Abonnement ID','Client Name', 'Version', 'Number of Months', 'Payment Date', 'Payment Method', 'Payment Status',
        'Start Date', 'End Date', 'Preavis Date', 'End Mail', 'End Mail Comment', 'End Viber', 'End Viber Comment',
        'Appointment Call', 'Appointment Call Comment'
    ])
    
    for abonnement in Abonnement.objects.all():
        client = abonnement.client
        renouvellement = Renouvellement.objects.filter(abonnement=abonnement).first()
        
        ws_abonnement.append([
            abonnement.id, f"{client.nom} {client.prenom}", abonnement.type_abonnement.version, abonnement.nombre_mois,
            abonnement.date_de_payement, abonnement.moyen_de_payement, abonnement.statut_de_payement,
            abonnement.date_debut_abonnement, renouvellement.date_fin_abonnement if renouvellement else '',
            renouvellement.date_preavis_renouvellement if renouvellement else '',
            renouvellement.mail_fin_abonnement if renouvellement else '', renouvellement.comment_mail_fin_abonnement if renouvellement else '',
            renouvellement.viber_fin_abonnement if renouvellement else '', renouvellement.comment_viber_fin_abonnement if renouvellement else '',
            renouvellement.appel_rdv_renouvellement if renouvellement else '', renouvellement.comment_appel_rdv_renouvellement if renouvellement else ''
        ])
    
    # Create the Boost worksheet
    ws_boost = workbook.create_sheet(title="Boost")
    ws_boost.append([
        'Abonnement ID', 'Month', 'FB Ad Publication', 'Price', 'Boost Date', 'End Date',
        'Client Service ID', 'Facebook Ad Mail', 'Facebook Ad Mail Comment', 'Facebook Result Mail', 
        'Facebook Result Mail Comment', 'Positive Testimonial Mail', 'Positive Testimonial Mail Comment', 
        'Result Call', 'Result Call Comment'
    ])
    
    for boost_service in BoostService.objects.all():
        client_service = ClientService.objects.filter(boost_service=boost_service).first()
        
        ws_boost.append([
            boost_service.abonnement.id, boost_service.mois, boost_service.publication_affiche_fb,
            boost_service.boost_prix, boost_service.date_boost, boost_service.date_fin,
            client_service.id if client_service else '',
            client_service.mail_pub_facebook if client_service else '', client_service.comment_mail_pub_facebook if client_service else '',
            client_service.mail_resultat_fb if client_service else '', client_service.comment_mail_resultat_fb if client_service else '',
            client_service.mail_temoignage_positive if client_service else '', client_service.comment_mail_temoignage_positive if client_service else '',
            client_service.appel_resultat if client_service else '', client_service.comment_appel_resultat if client_service else ''
        ])
    
    # Save workbook to an in-memory file
    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    
    # Create an HTTP response with the Excel file
    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=export.xlsx'
    return response


def import_from_excel(file):
    # Load the workbook and sheets
    workbook = openpyxl.load_workbook(filename=file)
    
    # Import Client data
    ws_client = workbook.get_sheet_by_name('Client')
    for row in ws_client.iter_rows(min_row=2, values_only=True):
        (
            client_or_prospect, nom, prenom, nom_commercial,  specialties, wilaya, commune,
            gps_insertion, gps_insertion_comment, photos_integration, photos_integration_comment,
            services_filling, services_filling_comment, client_remark, agenda, agenda_comment,
            contract_password, contract_password_comment, premium_search_logo, premium_search_logo_comment,
            ad_logo_order, welcome_mail, welcome_mail_comment, activation_call, activation_call_comment,
            telephones
        ) = row

        # Split phone numbers into a list
        phone_numbers = telephones.split(', ') if telephones else []

        # Create or update Client
        client, created = Client.objects.update_or_create(
            defaults={
                'client_or_prospect': client_or_prospect, 
                'nom_commercial': nom_commercial,
                'nom':nom, 
                'prenom':prenom, 
                'specialites': specialties,
                'wilaya': wilaya,
                'commune': commune
            }
        )

        # Create or update ClientProfil
        profil, created = ClientProfil.objects.update_or_create(
            client=client,
            defaults={
                'insertion_gps': gps_insertion,
                'comment_insertion_gps': gps_insertion_comment,
                'integration_photos': photos_integration,
                'comment_integration_photos': photos_integration_comment,
                'remplissage_services_cabinet': services_filling,
                'comment_remplissage_services_cabinet': services_filling_comment,
                'remarque_client': client_remark,
                'agenda': agenda,
                'comment_agenda': agenda_comment,
                'mot_de_passe_contrat': contract_password,
                'comment_mot_de_passe_contrat': contract_password_comment,
                'logo_recherche_premium': premium_search_logo,
                'comment_logo_recherche_premium': premium_search_logo_comment,
                'ordre_logo_pub_recherche': ad_logo_order,
                'mail_bienvenue': welcome_mail,
                'comment_mail_bienvenue': welcome_mail_comment,
                'appel_activation': activation_call,
                'comment_appel_activation': activation_call_comment
            }
        )

        # Add phone numbers to Client's Telephones
        for number in phone_numbers:
            if number:
                Telephone.objects.get_or_create(client=client, telephone_number=number)
    
    # Import AbonnementType data
    ws_abonnement_type = workbook.get_sheet_by_name('AbonnementType')
    for row in ws_abonnement_type.iter_rows(min_row=2, values_only=True):
        abonnement_type, price, version = row

        # Create or update AbonnementType
        abonnement_type_obj, created = AbonnementType.objects.update_or_create(
            abonnement=abonnement_type,
            defaults={'prix': price}
        )
        Version.objects.update_or_create(
            abonnement_type=abonnement_type_obj,
            version=version
        )
    
    # Import Abonnement data
    ws_abonnement = workbook.get_sheet_by_name('Abonnement')
    for row in ws_abonnement.iter_rows(min_row=2, values_only=True):
        (
            abonnement_id, client_name, version, number_of_months, payment_date, payment_method, payment_status,
            start_date, end_date, preavis_date, end_mail, end_mail_comment, end_viber, end_viber_comment,
            appointment_call, appointment_call_comment
        ) = row

        # Split client name
        client_names = client_name.split()
        client, _ = Client.objects.get_or_create(nom=client_names[0], prenom=client_names[1] if len(client_names) > 1 else '')

        # Create or update Abonnement
        abonnement, created = Abonnement.objects.update_or_create(
            id=abonnement_id,
            defaults={
                'client': client,
                'type_abonnement': AbonnementType.objects.filter(version=version).first(),
                'nombre_mois': number_of_months,
                'date_de_payement': payment_date,
                'moyen_de_payement': payment_method,
                'statut_de_payement': payment_status,
                'date_debut_abonnement': start_date
            }
        )
        Renouvellement.objects.update_or_create(
            abonnement=abonnement,
            defaults={
                'date_fin_abonnement': end_date,
                'date_preavis_renouvellement': preavis_date,
                'mail_fin_abonnement': end_mail,
                'comment_mail_fin_abonnement': end_mail_comment,
                'viber_fin_abonnement': end_viber,
                'comment_viber_fin_abonnement': end_viber_comment,
                'appel_rdv_renouvellement': appointment_call,
                'comment_appel_rdv_renouvellement': appointment_call_comment
            }
        )
    
    # Import Boost data
    ws_boost = workbook.get_sheet_by_name('Boost')
    for row in ws_boost.iter_rows(min_row=2, values_only=True):
        (
            abonnement_id, month, fb_ad_publication, price, boost_date, end_date,
            client_service_id, fb_ad_mail, fb_ad_mail_comment, fb_result_mail, fb_result_mail_comment,
            positive_testimonial_mail, positive_testimonial_mail_comment, result_call, result_call_comment
        ) = row

        abonnement = Abonnement.objects.filter(id=abonnement_id).first()
        client_service = ClientService.objects.filter(id=client_service_id).first()

        # Create or update BoostService
        BoostService.objects.update_or_create(
            abonnement=abonnement,
            mois=month,
            defaults={
                'publication_affiche_fb': fb_ad_publication,
                'boost_prix': price,
                'date_boost': boost_date,
                'date_fin': end_date
            }
        )
        if client_service:
            ClientService.objects.update_or_create(
                id=client_service_id,
                defaults={
                    'mail_pub_facebook': fb_ad_mail,
                    'comment_mail_pub_facebook': fb_ad_mail_comment,
                    'mail_resultat_fb': fb_result_mail,
                    'comment_mail_resultat_fb': fb_result_mail_comment,
                    'mail_temoignage_positive': positive_testimonial_mail,
                    'comment_mail_temoignage_positive': positive_testimonial_mail_comment,
                    'appel_resultat': result_call,
                    'comment_appel_resultat': result_call_comment
                }
            )
